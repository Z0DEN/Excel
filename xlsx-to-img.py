import openpyxl
from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import numpy as np
from openpyxl.utils import get_column_letter, column_index_from_string
import getpass

USERNAME = getpass.getuser()

img = Image.new('RGB', (100, 100), (255, 255, 255))
draw = ImageDraw.Draw(img)

workbook = openpyxl.load_workbook('C:/Users/'f'{USERNAME}''/Downloads/Excel-main/img-to-xlsx.xlsx')
sheets_list = workbook.sheetnames
sheet_active = workbook[sheets_list[0]]
x = 0
y = 0
for number in range(1, 101):
    cell_letter = get_column_letter(number)
    x += 1
    y = 0
    for cell_number in range(1, 101):
        cell_fill = str(sheet_active[f'{cell_letter}{cell_number}'].fill.start_color.index) #Получаем цвет ячейки
        cell_hex ='#' + cell_fill[2:8]
        draw.rectangle((x-1, y, x-1, y), fill=f'{cell_hex}')
        y += 1
img.show()
img.save('NewImg.jpg')
