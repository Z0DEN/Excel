from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
import getpass

USERNAME = getpass.getuser()

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")

im = Image.open('C:/Users/'f'{USERNAME}''/Downloads/Excel-main/test.jpg')
pixels = im.load() # список с пикселями
x, y = im.size # ширина (x) и высота (y) изображения

process = '#'
count = 0
deg = 0.05
degree = 5
space = '------------------------'
k = 1
g = 0
for i in range(x):
    for j in range(y):
        g += 1
        letter = get_column_letter(g)
        ws.column_dimensions[f'{letter}'].width = 3
        r, g, b = pixels[i, j]
        rgb = f"{r}, {g}, {b}"
        hex = (''.join([f"{int(i):02x}" for i in rgb.split(',')]))
        ws.cell(row=j+1, column=i+1).fill = PatternFill(start_color=f'{hex}', end_color=f'{hex}', fill_type="solid",)
        if ((i*j)/(x*y))>deg:
            print('process:','[',process,space[k:19],'] ---',degree,'%')
            deg += 0.05
            degree += 5
            process += '#'
            k += 1
wb.save('C:/Users/'f'{USERNAME}''/Downloads/Excel-main/img-to-xlsx.xlsx')
pro = '####################'
print('process:','[',pro,']','--- 100 %')